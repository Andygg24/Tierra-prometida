"""
grower_list.py — Genera Grower List preservando el molde original.
Usa openpyxl para editar el template conservando estilos, anchos y alturas.

JSON esperado por stdin:
{
  "importer":  "PRINCESSES KINGDOM CORP",
  "eta":       "22-06-26",
  "exporter":  "TIERRA PROMETIDA TRADING S.A.S.",
  "destino":   "MIAMI",
  "booking":   "ZIMUCRT914086",
  "container": "TTNU8272195",
  "growers": [
    { "registro":"15000896", "nombre":"Los Almendros",
      "dir":"Vereda Peñas", "ciudad":"Chocoita", "dpto":"Santander", "cajas":132 }
  ]
}
"""
import sys, json, io, os, tempfile
from copy import deepcopy
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

data      = json.load(io.TextIOWrapper(sys.stdin.buffer, encoding="utf-8-sig"))
importer  = data.get("importer",  "PRINCESSES KINGDOM CORP")
eta       = data.get("eta",       "")
exporter  = data.get("exporter",  "TIERRA PROMETIDA TRADING S.A.S.")
destino   = str(data.get("destino", "")).upper()
booking   = data.get("booking",   "")
container = data.get("container", "")
growers   = data.get("growers",   [])

BASE   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MOLDES = os.path.join(os.path.dirname(BASE), "Moldes")
TMPL   = os.path.join(MOLDES, "Grower List ZIMUCRT914086.xlsx")

wb = load_workbook(TMPL)
ws = wb.active

PACKING_FACILITY = "C.I. GRAMALUZ CRA 35 A 46-04 BUCARAMANGA"
DATA_START = 3   # primera fila de datos (fila 1=email, fila 2=headers)
N_COLS     = 17  # columnas A-Q

# ── Capturar estilos de la fila de datos de referencia (fila 3) ───────
ref_styles = {}
for col_idx in range(1, N_COLS + 1):
    cell = ws.cell(row=DATA_START, column=col_idx)
    if not isinstance(cell, MergedCell) and hasattr(cell, "_style"):
        ref_styles[col_idx] = deepcopy(cell._style)

# ── Limpiar todas las filas de datos existentes ────────────────────────
for r in range(DATA_START, ws.max_row + 1):
    for col_idx in range(1, N_COLS + 1):
        ws.cell(row=r, column=col_idx).value = None

# ── Escribir una fila por predio ───────────────────────────────────────
total_cajas = 0
for i, g in enumerate(growers):
    r     = DATA_START + i
    cajas = int(g.get("cajas", 0))
    total_cajas += cajas

    # Aplicar estilo de la fila de referencia
    for col_idx in range(1, N_COLS + 1):
        cell = ws.cell(row=r, column=col_idx)
        st = ref_styles.get(col_idx)
        if st is not None:
            cell._style = deepcopy(st)

    ws.cell(row=r, column=1).value  = importer
    ws.cell(row=r, column=2).value  = eta
    ws.cell(row=r, column=3).value  = exporter
    ws.cell(row=r, column=4).value  = destino
    ws.cell(row=r, column=5).value  = g.get("registro", "")
    ws.cell(row=r, column=6).value  = g.get("nombre",   "")
    ws.cell(row=r, column=7).value  = g.get("dir",      "")
    ws.cell(row=r, column=8).value  = g.get("ciudad",   "")
    ws.cell(row=r, column=9).value  = g.get("dpto",     "")
    ws.cell(row=r, column=10).value = PACKING_FACILITY
    ws.cell(row=r, column=11).value = None   # FDA reg — en blanco
    ws.cell(row=r, column=12).value = "LIMON TAHITI"
    ws.cell(row=r, column=13).value = cajas
    ws.cell(row=r, column=14).value = "CAJAS"
    ws.cell(row=r, column=15).value = 16.2
    ws.cell(row=r, column=16).value = booking
    ws.cell(row=r, column=17).value = container

# ── Fila de total (columna M) ──────────────────────────────────────────
total_row = DATA_START + len(growers)
for col_idx in range(1, N_COLS + 1):
    cell = ws.cell(row=total_row, column=col_idx)
    st = ref_styles.get(col_idx)
    if st is not None:
        cell._style = deepcopy(st)
ws.cell(row=total_row, column=13).value = total_cajas

print(f"Predios: {len(growers)} | Total cajas: {total_cajas}", file=sys.stderr)

# ── Guardar y enviar por stdout ───────────────────────────────────────
tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
tmp_path = tmp.name
tmp.close()

try:
    wb.save(tmp_path)
    with open(tmp_path, "rb") as f:
        sys.stdout.buffer.write(f.read())
finally:
    try:
        os.unlink(tmp_path)
    except Exception:
        pass
