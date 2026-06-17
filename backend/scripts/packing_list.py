"""
packing_list.py — Genera Packing List preservando el molde original.
Estrategia:
  1. Capturar estilos de TODAS las filas de datos ANTES de cualquier cambio
     (incluyendo estilos de celdas que son MergedCell no-ancla).
  2. Desunir merges del área de datos.
  3. Limpiar valores.
  4. Escribir nuevos valores.
  5. Re-aplicar los estilos capturados (garantiza que el formato del molde
     se preserve en filas existentes Y en filas extra para pallets mixtos).

JSON esperado por stdin:
{
  "plNo":              "2026-174",
  "destino":           "Philadelphia",
  "fechaCargue":       "2026-05-15",
  "empresaTransporte": "Transportando Express",
  "placa":             "QJN678",
  "tempRecorder":      "V1-0041573",
  "finalStamps":       "005743-SQ83066",
  "totalCajas":        1400,
  "pallets": [
    {"id": 1,  "calibres": [{"size": 200, "cajas": 70, "predio": "Las Brisas", "ica": "980005905"}]},
    {"id": 15, "calibres": [
        {"size": 110, "cajas": 60, "predio": "La Esperanza", "ica": "430003503"},
        {"size": 230, "cajas": 10, "predio": "San Pedro",    "ica": "590004304"}
    ]}
  ]
}
"""
import sys, json, os, io, zipfile, re
from copy import deepcopy
from datetime import date as date_cls
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string

# ── Leer JSON de stdin ────────────────────────────────────────────────────
data = json.load(io.TextIOWrapper(sys.stdin.buffer, encoding="utf-8-sig"))

pl_no        = str(data.get("plNo", ""))
destino      = str(data.get("destino", "PHILADELPHIA")).upper()
fecha_str    = data.get("fechaCargue", "")
empresa      = str(data.get("empresaTransporte", ""))
placa        = str(data.get("placa", ""))
temp_rec     = str(data.get("tempRecorder", ""))
final_stamps = str(data.get("finalStamps", ""))
total_cajas  = int(data.get("totalCajas", 1400))
pallets      = data.get("pallets", [])

PESO_KG    = 16.2
DATA_START = 11
COLS       = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
COL_IDXS   = {c: column_index_from_string(c) for c in COLS}

BASE   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MOLDES = os.path.join(os.path.dirname(BASE), "Moldes")
TMPL   = os.path.join(MOLDES, "Packing List 271169365.xlsx")

# ── Cargar molde ──────────────────────────────────────────────────────────
wb = load_workbook(TMPL, keep_vba=False)
ws = wb.active

TMPL_LAST_ROW = ws.max_row   # 31 en el molde actual

try:
    n_img = len(ws._images)
except Exception:
    n_img = "N/A"
print(f"Imágenes en el workbook: {n_img}", file=sys.stderr)
print(f"AutoFilter: {ws.auto_filter.ref}", file=sys.stderr)
print(f"Última fila del molde: {TMPL_LAST_ROW}", file=sys.stderr)

# ═══════════════════════════════════════════════════════════════════════════
# PASO 1 — Capturar estilos y alturas ANTES de cualquier modificación
# Para MergedCells no-ancla, usamos el estilo de su celda ancla.
# ═══════════════════════════════════════════════════════════════════════════

# Mapa (row, col_idx) -> estilo del ancla, para celdas dentro de un merge
anchor_style_map = {}
for mr in ws.merged_cells.ranges:
    if mr.min_row >= DATA_START:
        anchor = ws.cell(row=mr.min_row, column=mr.min_col)
        anchor_st = deepcopy(anchor._style) if hasattr(anchor, "_style") else None
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                anchor_style_map[(r, c)] = anchor_st

# Capturar estilo de cada celda de datos
captured = {}   # row -> {col: _style, "_height": float|None}
for r in range(DATA_START, TMPL_LAST_ROW + 1):
    captured[r] = {}
    for col in COLS:
        cidx = COL_IDXS[col]
        cell = ws.cell(row=r, column=cidx)
        if isinstance(cell, MergedCell):
            st = anchor_style_map.get((r, cidx))
        else:
            st = deepcopy(cell._style) if hasattr(cell, "_style") else None
        captured[r][col] = st
    captured[r]["_height"] = ws.row_dimensions[r].height

# Fila de referencia para filas extra más allá del molde (fila intermedia)
REF_ROW = 15

# ═══════════════════════════════════════════════════════════════════════════
# PASO 2 — Escribir encabezado (solo .value, los estilos ya están en molde)
# ═══════════════════════════════════════════════════════════════════════════
peso = round(total_cajas * PESO_KG, 2)

ws["A2"].value = (
    "TIERRA PROMETIDA TRADING SAS.\n"
    "BARRANQUILLA - ATLANTICO\n"
    "operaciones@tierraprometidat.com\n"
    f"Packing List No. {pl_no}"
)
ws["D5"].value = destino
ws["F5"].value = total_cajas
ws["H5"].value = len(pallets)
ws["A7"].value = peso
ws["D7"].value = peso
ws["F7"].value = empresa
ws["H7"].value = placa
ws["A9"].value = pl_no
ws["D9"].value = temp_rec
ws["H9"].value = final_stamps

if fecha_str:
    try:
        y, mo, d = [int(x) for x in fecha_str.split("-")]
        ws["F9"].value = date_cls(y, mo, d)
    except Exception:
        ws["F9"].value = fecha_str

# ═══════════════════════════════════════════════════════════════════════════
# PASO 3 — Construir lista plana de filas de pallets
# ═══════════════════════════════════════════════════════════════════════════
sorted_pallets = sorted(pallets, key=lambda p: p.get("id", 0))
rows = []
merges_to_add = []

for p in sorted_pallets:
    pid      = p.get("id", 0)
    calibres = p.get("calibres") or [{"size": "", "cajas": 0, "predio": "", "ica": ""}]
    r_start  = DATA_START + len(rows)

    for ci, c in enumerate(calibres):
        rows.append({
            "pid":     pid if ci == 0 else None,
            "size":    c.get("size", ""),
            "cajas":   int(c.get("cajas", 0)),
            "predio":  c.get("predio", ""),
            "ica":     str(c.get("ica", "")),
            "is_cont": ci > 0,
        })

    if len(calibres) > 1:
        r_end = DATA_START + len(rows) - 1
        for col in ("A", "B"):
            merges_to_add.append((col, r_start, r_end))

# ═══════════════════════════════════════════════════════════════════════════
# PASO 4 — Desunir merges del área de datos (ahora seguro, estilos guardados)
# ═══════════════════════════════════════════════════════════════════════════
merges_data = [str(m) for m in ws.merged_cells.ranges if m.min_row >= DATA_START]
for ref in merges_data:
    ws.unmerge_cells(ref)
print(f"Merges desunidos: {len(merges_data)}", file=sys.stderr)

# ═══════════════════════════════════════════════════════════════════════════
# PASO 5 — Limpiar valores del área de datos
# ═══════════════════════════════════════════════════════════════════════════
for r in range(DATA_START, TMPL_LAST_ROW + 1):
    for col in COLS:
        ws.cell(row=r, column=COL_IDXS[col]).value = None

# ═══════════════════════════════════════════════════════════════════════════
# PASO 6 — Escribir nuevos datos y re-aplicar estilos capturados
# ═══════════════════════════════════════════════════════════════════════════
def style_for(excel_row, col):
    """Devuelve el estilo capturado del molde para esta fila/columna."""
    src_row = excel_row if excel_row <= TMPL_LAST_ROW else REF_ROW
    return captured.get(src_row, {}).get(col)

def height_for(excel_row):
    src_row = excel_row if excel_row <= TMPL_LAST_ROW else REF_ROW
    return captured.get(src_row, {}).get("_height")

for i, rd in enumerate(rows):
    excel_row = DATA_START + i

    # Re-aplicar estilo capturado a cada celda de esta fila
    for col in COLS:
        cell = ws.cell(row=excel_row, column=COL_IDXS[col])
        st = style_for(excel_row, col)
        if st is not None:
            cell._style = deepcopy(st)

    # Re-aplicar altura de fila
    ht = height_for(excel_row)
    if ht:
        ws.row_dimensions[excel_row].height = ht

    # Escribir valores
    if not rd["is_cont"]:
        ws[f"A{excel_row}"].value = rd["pid"]
        ws[f"B{excel_row}"].value = "16.2 KG"

    ws[f"C{excel_row}"].value = rd["size"] if rd["size"] != "" else None
    ws[f"D{excel_row}"].value = "LIMON TAHITI"
    ws[f"E{excel_row}"].value = 1
    ws[f"F{excel_row}"].value = rd["cajas"]
    ws[f"G{excel_row}"].value = rd["predio"] or None
    ws[f"H{excel_row}"].value = rd["cajas"]
    ws[f"I{excel_row}"].value = rd["ica"] or None

# ═══════════════════════════════════════════════════════════════════════════
# PASO 7 — Agregar merges para pallets mixtos
# ═══════════════════════════════════════════════════════════════════════════
for col, r1, r2 in merges_to_add:
    try:
        ws.merge_cells(f"{col}{r1}:{col}{r2}")
    except Exception as e:
        print(f"Merge {col}{r1}:{col}{r2} omitido: {e}", file=sys.stderr)

# ── Diagnóstico final ─────────────────────────────────────────────────────
try:
    n_img_post = len(ws._images)
except Exception:
    n_img_post = "N/A"
print(f"Imágenes (post): {n_img_post}", file=sys.stderr)
print(f"AutoFilter (post): {ws.auto_filter.ref}", file=sys.stderr)
extra = max(0, len(rows) - (TMPL_LAST_ROW - DATA_START + 1))
print(f"Pallets: {len(sorted_pallets)} | Filas: {len(rows)} | Extra: {extra}", file=sys.stderr)

# ── Guardar + restaurar logo/drawing (openpyxl los descarta al guardar) ───
opxl_buf = io.BytesIO()
wb.save(opxl_buf)
opxl_buf.seek(0)

# Archivos de imagen/dibujo que openpyxl siempre elimina
DRAWING_FILES = [
    "xl/drawings/drawing1.xml",
    "xl/drawings/_rels/drawing1.xml.rels",
    "xl/media/image1.png",
]

# Fragmentos a inyectar en [Content_Types].xml
CT_PNG  = '<Default Extension="png" ContentType="image/png"/>'
CT_DRW  = '<Override PartName="/xl/drawings/drawing1.xml" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'

# Relación del dibujo a inyectar en sheet1.xml.rels
REL_DRW = '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/>'

with zipfile.ZipFile(TMPL, "r") as tmpl_zip:
    tmpl_files = {n: tmpl_zip.read(n) for n in DRAWING_FILES if n in tmpl_zip.namelist()}

out_buf = io.BytesIO()
with zipfile.ZipFile(opxl_buf, "r") as src_zip, \
     zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as dst_zip:

    for item in src_zip.infolist():
        data = src_zip.read(item.filename)

        if item.filename == "[Content_Types].xml":
            text = data.decode("utf-8")
            # Agregar png si no está
            if 'Extension="png"' not in text:
                text = text.replace('<Default Extension="rels"', f'{CT_PNG}<Default Extension="rels"', 1)
            # Agregar drawing override si no está
            if "drawings/drawing1.xml" not in text:
                text = text.replace("</Types>", f"{CT_DRW}</Types>", 1)
            data = text.encode("utf-8")

        elif item.filename == "xl/worksheets/_rels/sheet1.xml.rels":
            text = data.decode("utf-8")
            # Inyectar rId3 de drawing si no está
            if "drawing1.xml" not in text:
                text = text.replace("</Relationships>", f"{REL_DRW}</Relationships>", 1)
            data = text.encode("utf-8")

        dst_zip.writestr(item, data)

    # Copiar archivos de dibujo/imagen desde el molde
    for path, fdata in tmpl_files.items():
        dst_zip.writestr(path, fdata)

out_buf.seek(0)
sys.stdout.buffer.write(out_buf.read())
